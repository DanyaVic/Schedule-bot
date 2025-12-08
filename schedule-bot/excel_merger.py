import openpyxl
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelMerger:
    """Развёртывает объединённые ячейки в новый файл"""

    def __init__(self, input_file: str, output_file: str):
        self.input_file = input_file
        self.output_file = output_file

    def merge_cells(self) -> bool:
        try:
            logger.info(f"📂 Читаем файл: {self.input_file}")
            wb = openpyxl.load_workbook(self.input_file)
            ws = wb.active

            # Создаём новый workbook и лист
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active

            # Копируем всё содержимое как есть сначала
            max_row = ws.max_row
            max_col = ws.max_column
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    new_ws.cell(row=r, column=c).value = ws.cell(row=r, column=c).value

            logger.info("🔍 Ищем объединённые диапазоны...")
            merged_ranges = list(ws.merged_cells.ranges)
            logger.info(f"✅ Найдено объединённых диапазонов: {len(merged_ranges)}")

            # Теперь разворачиваем объединённые диапазоны в НОВОМ листе
            for merged_range in merged_ranges:
                self._expand_merged_range(ws, new_ws, merged_range)

            logger.info(f"💾 Сохраняем в: {self.output_file}")
            new_wb.save(self.output_file)
            logger.info("✅ Готово!")
            return True

        except Exception as e:
            logger.error(f"❌ Ошибка: {e}")
            return False

    def _expand_merged_range(self, src_ws, dst_ws, merged_range):
        """
        src_ws — исходный лист с настоящими merged-ячейками
        dst_ws — новый лист, где мы просто дублируем значения
        """
        range_str = str(merged_range)
        start_cell, end_cell = range_str.split(':')

        start_col, start_row = self._parse_cell(start_cell)
        end_col, end_row = self._parse_cell(end_cell)

        # значение из верхней левой ячейки ИСТОЧНИКА
        top_left_val = src_ws.cell(row=start_row, column=start_col).value
        logger.info(f"  Диапазон {range_str}: значение '{top_left_val}'")

        # Во ВСЕ ячейки диапазона НОВОГО листа записываем это значение
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                dst_ws.cell(row=r, column=c).value = top_left_val

    @staticmethod
    def _parse_cell(cell_str: str) -> tuple:
        col_str = ""
        row_str = ""
        for ch in cell_str:
            if ch.isalpha():
                col_str += ch
            else:
                row_str += ch
        col = openpyxl.utils.column_index_from_string(col_str)
        row = int(row_str)
        return col, row


if __name__ == "__main__":
    merger = ExcelMerger(
        input_file='schedules/schedule.xlsx',
        output_file='schedules/schedule_merged.xlsx'
    )
    if merger.merge_cells():
        print("\n✅ Файл успешно развёрнут!")
        print("📄 Используйте 'schedules/schedule_merged.xlsx' в парсере")
    else:
        print("\n❌ Ошибка при развёртывании файла")
